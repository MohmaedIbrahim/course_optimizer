import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import pulp
import numpy as np
from typing import Dict, List, Tuple
import io
from scipy.cluster.hierarchy import dendrogram, linkage
from scipy.spatial.distance import pdist

class CourseCoveringProblem:
    """Course covering optimization problem solver matching exact mathematical formulation."""
    
    def __init__(self, courses: List[str], professors: List[str], terms: List[str],
                 course_preferences: Dict[Tuple[str, str], float],
                 term_preferences: Dict[Tuple[str, str], float],
                 course_streams: Dict[Tuple[str, str], int],
                 professor_total_load: Dict[str, int],
                 professor_term_limits: Dict[Tuple[str, str], int],
                 course_offerings: Dict[Tuple[str, str], int]):
        
        self.courses = courses
        self.professors = professors
        self.terms = terms
        self.course_preferences = course_preferences
        self.term_preferences = term_preferences
        self.course_streams = course_streams
        self.professor_total_load = professor_total_load
        self.professor_term_limits = professor_term_limits
        self.course_offerings = course_offerings
        
        self.model = None
        self.x_vars = {}
        
        self._validate_preferences()
        
    def _validate_preferences(self):
        """Validate that preference scores are in 0-10 range."""
        for (course, prof), score in self.course_preferences.items():
            if not (0 <= score <= 10):
                raise ValueError(f"Course preference c_{{{course},{prof}}} = {score} must be in range [0,10]")
        
        for (prof, term), score in self.term_preferences.items():
            if not (0 <= score <= 10):
                raise ValueError(f"Term preference t_{{{prof},{term}}} = {score} must be in range [0,10]")
    
    def build_model(self):
        """Build the optimization model matching the mathematical formulation exactly."""
        self.model = pulp.LpProblem("Course_Covering_Mathematical", pulp.LpMaximize)
        
        self.x_vars = {}
        for course in self.courses:
            for professor in self.professors:
                for term in self.terms:
                    if self.course_offerings.get((course, term), 0) == 1:
                        self.x_vars[(course, professor, term)] = pulp.LpVariable(
                            f"x_{course}_{professor}_{term}", cat='Binary'
                        )
        
        course_pref_term = pulp.lpSum([
            self.course_preferences.get((course, professor), 0) * self.x_vars[(course, professor, term)]
            for course in self.courses
            for professor in self.professors
            for term in self.terms
            if (course, professor, term) in self.x_vars
        ])
        
        term_pref_term = pulp.lpSum([
            self.term_preferences.get((professor, term), 0) * self.x_vars[(course, professor, term)]
            for course in self.courses
            for professor in self.professors
            for term in self.terms
            if (course, professor, term) in self.x_vars
        ])
        
        self.model += course_pref_term + term_pref_term
        self._add_constraints()
        
    def _add_constraints(self):
        """Add all constraints matching the mathematical formulation exactly."""
        
        for professor in self.professors:
            for term in self.terms:
                term_stream_load = pulp.lpSum([
                    self.course_streams.get((course, term), 0) * self.x_vars[(course, professor, term)]
                    for course in self.courses
                    if (course, professor, term) in self.x_vars
                ])
                max_streams_in_term = self.professor_term_limits.get((professor, term), 0)
                self.model += (
                    term_stream_load <= max_streams_in_term,
                    f"StreamLoad_L_{professor}_{term}"
                )
        
        for professor in self.professors:
            total_courses_assigned = pulp.lpSum([
                self.x_vars[(course, professor, term)]
                for course in self.courses
                for term in self.terms
                if (course, professor, term) in self.x_vars
            ])
            self.model += (
                total_courses_assigned <= self.professor_total_load[professor],
                f"TotalLoad_b_{professor}"
            )
        
        for course in self.courses:
            for term in self.terms:
                if self.course_offerings.get((course, term), 0) == 1:
                    self.model += (
                        pulp.lpSum([
                            self.x_vars[(course, professor, term)]
                            for professor in self.professors
                            if (course, professor, term) in self.x_vars
                        ]) == 1,
                        f"CourseOffering_O_{course}_{term}"
                    )
    
    def solve(self):
        """Solve the optimization problem."""
        if self.model is None:
            self.build_model()
        
        solver = pulp.PULP_CBC_CMD(msg=0)
        self.model.solve(solver)
        
        status = pulp.LpStatus[self.model.status]
        
        if status == 'Optimal':
            return self._extract_solution()
        else:
            return {
                'status': status,
                'objective_value': None,
                'assignments': {},
                'professor_loads': {},
                'unassigned_offerings': [],
                'constraint_violations': self._analyze_infeasibility() if status == 'Infeasible' else None
            }
    
    def _extract_solution(self):
        """Extract solution from solved model."""
        assignments = {}
        professor_loads = {
            prof: {
                'total_courses': 0,
                'streams_per_term': {term: 0 for term in self.terms},
                'total_streams': 0
            } 
            for prof in self.professors
        }
        unassigned_offerings = []
        
        for course in self.courses:
            for term in self.terms:
                if self.course_offerings.get((course, term), 0) == 1:
                    assigned = False
                    for professor in self.professors:
                        if (course, professor, term) in self.x_vars and self.x_vars[(course, professor, term)].varValue == 1:
                            assignments[(course, term)] = professor
                            professor_loads[professor]['total_courses'] += 1
                            streams_count = self.course_streams.get((course, term), 1)
                            professor_loads[professor]['streams_per_term'][term] += streams_count
                            professor_loads[professor]['total_streams'] += streams_count
                            assigned = True
                            break
                    
                    if not assigned:
                        unassigned_offerings.append((course, term))
        
        return {
            'status': 'Optimal',
            'objective_value': pulp.value(self.model.objective),
            'assignments': assignments,
            'professor_loads': professor_loads,
            'unassigned_offerings': unassigned_offerings
        }
    
    def _analyze_infeasibility(self):
        """Analyze potential constraint violations if problem is infeasible."""
        violations = []
        
        total_required_streams = sum([
            self.course_streams.get((course, term), 1)
            for course in self.courses
            for term in self.terms
            if self.course_offerings.get((course, term), 0) == 1
        ])
        
        total_available_streams = sum([
            self.professor_term_limits.get((prof, term), 0)
            for prof in self.professors
            for term in self.terms
        ])
        
        if total_required_streams > total_available_streams:
            violations.append(f"Total required streams ({total_required_streams}) exceed total available capacity ({total_available_streams})")
        
        total_required_courses = sum([
            1 for course in self.courses
            for term in self.terms
            if self.course_offerings.get((course, term), 0) == 1
        ])
        
        total_available_courses = sum(self.professor_total_load.values())
        
        if total_required_courses > total_available_courses:
            violations.append(f"Total required courses ({total_required_courses}) exceed total available slots ({total_available_courses})")
        
        return violations


def main():
    st.set_page_config(
        page_title="RASTA-OP",
        page_icon="🎓",
        layout="wide"
    )
    
    st.title("RASTA-OP")
    st.markdown("Optimize faculty assignments using exact mathematical formulation with L_jk and b_j constraints")
    st.markdown("---")
    
    input_method = st.sidebar.selectbox(
        "Choose Input Method:",
        ["Manual Input", "Excel Upload"]
    )
    
    if 'step' not in st.session_state:
        st.session_state.step = 1
    if 'courses' not in st.session_state:
        st.session_state.courses = []
    if 'professors' not in st.session_state:
        st.session_state.professors = []
    if 'terms' not in st.session_state:
        st.session_state.terms = ['T1', 'T2', 'T3']
    
    if input_method == "Excel Upload":
        if st.session_state.step == 1:
            show_excel_upload_step()
        elif st.session_state.step == 2:
            show_data_analysis_step()
        elif st.session_state.step == 3:
            show_results_step()
    else:
        if st.session_state.step == 1:
            show_setup_step()
        elif st.session_state.step == 2:
            show_constraints_step()
        elif st.session_state.step == 3:
            show_preferences_step()
        elif st.session_state.step == 4:
            show_results_step()


def show_setup_step():
    """Show the setup step for courses and professors."""
    st.header("Step 1: Setup Courses and Professors")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("Courses")
        courses_input = st.text_area(
            "Enter courses (one per line):",
            value="ACTL1\nACTL2\nACTL3\nACTL4\nACTL5",
            height=120
        )
    
    with col2:
        st.subheader("Professors")
        professors_input = st.text_area(
            "Enter professors (one per line):",
            value="Jonathan\nJK\nPatrick\nAndres",
            height=120
        )
    
    if st.button("Next: Set Constraints", type="primary"):
        courses = [course.strip() for course in courses_input.split('\n') if course.strip()]
        professors = [prof.strip() for prof in professors_input.split('\n') if prof.strip()]
        
        if not courses or not professors:
            st.error("Please enter at least one course and one professor.")
        else:
            st.session_state.courses = courses
            st.session_state.professors = professors
            st.session_state.step = 2
            st.rerun()


def show_constraints_step():
    """Show constraints configuration step."""
    st.header("Step 2: Configure Constraints")
    
    courses = st.session_state.courses
    professors = st.session_state.professors
    terms = st.session_state.terms
    
    st.subheader("Course Offerings & Streams (O_ik, n_ik)")
    st.markdown("Configure which courses are offered in which terms and how many streams each has")
    
    course_offerings = {}
    course_streams = {}
    
    for course in courses:
        st.write(f"**{course}**")
        cols = st.columns(len(terms))
        
        for idx, term in enumerate(terms):
            with cols[idx]:
                offered = st.checkbox(f"Offer in {term}", key=f"offer_{course}_{term}")
                course_offerings[(course, term)] = 1 if offered else 0
                
                if offered:
                    streams = st.number_input(
                        f"Streams in {term}:",
                        min_value=1, max_value=5, value=1,
                        key=f"streams_{course}_{term}"
                    )
                    course_streams[(course, term)] = streams
                else:
                    course_streams[(course, term)] = 0
    
    st.subheader("Professor Constraints")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**Total Teaching Load (b_j)**")
        st.markdown("Maximum total courses per professor across all terms")
        professor_total_load = {}
        
        for professor in professors:
            load = st.number_input(
                f"{professor} - Total courses (b_j):",
                min_value=1, max_value=10, value=4,
                key=f"total_load_{professor}"
            )
            professor_total_load[professor] = load
    
    with col2:
        st.write("**Term Stream Limits (L_jk)**")
        st.markdown("Maximum streams per professor per term (independent from b_j)")
        professor_term_limits = {}
        
        for professor in professors:
            st.write(f"**{professor}:**")
            sub_cols = st.columns(len(terms))
            
            for idx, term in enumerate(terms):
                with sub_cols[idx]:
                    limit = st.number_input(
                        f"{term}:",
                        min_value=0, max_value=6, value=2,
                        key=f"term_limit_{professor}_{term}",
                        help=f"Max streams {professor} can teach in {term}"
                    )
                    professor_term_limits[(professor, term)] = limit
    
    st.subheader("Constraint Summary")
    
    st.write("**Course Offerings Matrix (O_ik):**")
    offerings_data = []
    for course in courses:
        row = {'Course': course}
        for term in terms:
            if course_offerings.get((course, term), 0) == 1:
                streams = course_streams.get((course, term), 1)
                row[term] = f"✓ ({streams} streams)"
            else:
                row[term] = "✗"
        offerings_data.append(row)
    
    offerings_df = pd.DataFrame(offerings_data)
    st.dataframe(offerings_df, hide_index=True)
    
    st.write("**Professor Constraints Summary:**")
    prof_summary = []
    for prof in professors:
        row = {'Professor': prof, 'Total Load (b_j)': professor_total_load[prof]}
        for term in terms:
            row[f"{term} Limit (L_jk)"] = professor_term_limits[(prof, term)]
        prof_summary.append(row)
    
    prof_df = pd.DataFrame(prof_summary)
    st.dataframe(prof_df, hide_index=True)
    
    st.session_state.course_offerings = course_offerings
    st.session_state.course_streams = course_streams
    st.session_state.professor_total_load = professor_total_load
    st.session_state.professor_term_limits = professor_term_limits
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("← Back to Setup"):
            st.session_state.step = 1
            st.rerun()
    
    with col2:
        if st.button("Next: Set Preferences", type="primary"):
            st.session_state.step = 3
            st.rerun()


def show_preferences_step():
    """Show preferences configuration step."""
    st.header("Step 3: Set Preferences (0-10 Scale)")
    
    courses = st.session_state.courses
    professors = st.session_state.professors
    terms = st.session_state.terms
    
    st.subheader("Course Preferences (c_ij)")
    st.markdown("**Scale: 0 = Cannot teach, 5 = Neutral, 10 = Strongly prefer**")
    
    course_preferences = {}
    
    for professor in professors:
        st.write(f"**{professor}'s Course Preferences:**")
        cols = st.columns(min(4, len(courses)))
        
        for idx, course in enumerate(courses):
            with cols[idx % len(cols)]:
                pref = st.slider(
                    f"{course}",
                    min_value=0, max_value=10, value=5,
                    key=f"course_pref_{course}_{professor}"
                )
                course_preferences[(course, professor)] = pref
    
    st.subheader("Term Preferences (t_jk)")
    st.markdown("**Scale: 0 = Cannot teach, 5 = Neutral, 10 = Strongly prefer**")
    
    term_preferences = {}
    
    for professor in professors:
        st.write(f"**{professor}'s Term Preferences:**")
        cols = st.columns(len(terms))
        
        for idx, term in enumerate(terms):
            with cols[idx]:
                pref = st.slider(
                    f"{term}",
                    min_value=0, max_value=10, value=5,
                    key=f"term_pref_{professor}_{term}"
                )
                term_preferences[(professor, term)] = pref
    
    st.subheader("Preference Matrices")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**Course Preferences (c_ij):**")
        course_pref_matrix = pd.DataFrame(index=courses, columns=professors)
        for course in courses:
            for prof in professors:
                course_pref_matrix.loc[course, prof] = course_preferences.get((course, prof), 0)
        st.dataframe(course_pref_matrix)
    
    with col2:
        st.write("**Term Preferences (t_jk):**")
        term_pref_matrix = pd.DataFrame(index=professors, columns=terms)
        for prof in professors:
            for term in terms:
                term_pref_matrix.loc[prof, term] = term_preferences.get((prof, term), 0)
        st.dataframe(term_pref_matrix)
    
    st.session_state.course_preferences = course_preferences
    st.session_state.term_preferences = term_preferences
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("← Back to Constraints"):
            st.session_state.step = 2
            st.rerun()
    
    with col2:
        if st.button("Run Optimization", type="primary"):
            st.session_state.step = 4
            st.rerun()


def show_results_step():
    """Show optimization results."""
    st.header("Optimization Results")
    
    courses = st.session_state.courses
    professors = st.session_state.professors
    terms = st.session_state.terms
    course_offerings = st.session_state.course_offerings
    course_streams = st.session_state.course_streams
    professor_total_load = st.session_state.professor_total_load
    professor_term_limits = st.session_state.professor_term_limits
    course_preferences = st.session_state.course_preferences
    term_preferences = st.session_state.term_preferences
    
    with st.spinner("Running optimization with mathematical formulation..."):
        try:
            problem = CourseCoveringProblem(
                courses=courses,
                professors=professors,
                terms=terms,
                course_preferences=course_preferences,
                term_preferences=term_preferences,
                course_streams=course_streams,
                professor_total_load=professor_total_load,
                professor_term_limits=professor_term_limits,
                course_offerings=course_offerings
            )
            
            solution = problem.solve()
            
        except ValueError as e:
            st.error(f"Invalid preferences: {str(e)}")
            return
        except Exception as e:
            st.error(f"Optimization error: {str(e)}")
            return
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Status", solution['status'])
    with col2:
        if solution['objective_value']:
            st.metric("Objective Value", f"{solution['objective_value']:.1f}")
        else:
            st.metric("Objective Value", "N/A")
    with col3:
        unassigned_count = len(solution.get('unassigned_offerings', []))
        st.metric("Unassigned Offerings", unassigned_count)
    
    if solution['status'] == 'Optimal':
        st.success("Optimization completed successfully!")
        
        assignments = solution.get('assignments', {})
        
        st.subheader("Course Assignment Matrix - All Terms")
        st.markdown("**Rows = Courses, Columns = Staff Members, Values = Term (T1, T2, T3) or 0 (Not Assigned)**")
        
        matrix_data = []
        for course in courses:
            row_data = {'Course': course}
            for professor in professors:
                row_data[professor] = 0
            
            for (assigned_course, term), assigned_professor in assignments.items():
                if assigned_course == course:
                    row_data[assigned_professor] = term
            
            matrix_data.append(row_data)
        
        main_matrix_df = pd.DataFrame(matrix_data)
        main_matrix_df.set_index('Course', inplace=True)
        
        st.dataframe(
            main_matrix_df,
            height=min(600, len(courses) * 25),
            use_container_width=True
        )
        
        total_assignments = (main_matrix_df != 0).sum().sum()
        courses_with_staff = (main_matrix_df != 0).any(axis=1).sum()
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Total Assignments", int(total_assignments))
        with col2:
            st.metric("Courses Covered", f"{courses_with_staff}/{len(courses)}")
        
        term_counts = {}
        for term in ['T1', 'T2', 'T3']:
            count = (main_matrix_df == term).sum().sum()
            term_counts[term] = count
        
        col3, col4, col5 = st.columns(3)
        with col3:
            st.metric("T1 Assignments", term_counts["T1"])
        with col4:
            st.metric("T2 Assignments", term_counts["T2"])
        with col5:
            st.metric("T3 Assignments", term_counts["T3"])
        
        st.subheader("Term-Specific Assignment Matrices")
        
        term_tabs = st.tabs([f"Term {term}" for term in terms])
        
        for tab_idx, term in enumerate(terms):
            with term_tabs[tab_idx]:
                st.markdown(f"**{term} Matrix: Rows = Courses, Columns = Staff**")
                
                term_matrix_data = []
                for course in courses:
                    term_row = {'Course': course}
                    for professor in professors:
                        term_row[professor] = 0
                    
                    for (assigned_course, assigned_term), assigned_professor in assignments.items():
                        if assigned_course == course and assigned_term == term:
                            term_row[assigned_professor] = 1
                    
                    term_matrix_data.append(term_row)
                
                term_matrix_df = pd.DataFrame(term_matrix_data)
                term_matrix_df.set_index('Course', inplace=True)
                
                st.dataframe(
                    term_matrix_df,
                    height=min(500, len(courses) * 22),
                    use_container_width=True
                )
                
                term_total = term_matrix_df.sum().sum()
                term_courses = (term_matrix_df.sum(axis=1) > 0).sum()
                
                col1, col2 = st.columns(2)
                with col1:
                    st.metric(f"{term} Assignments", int(term_total))
                with col2:
                    st.metric(f"{term} Courses", int(term_courses))
        
        st.subheader("Professor Workload Analysis")
        
        workload_data = []
        for professor in professors:
            prof_load = solution['professor_loads'][professor]
            
            total_courses = prof_load['total_courses']
            max_total = professor_total_load[professor]
            
            workload_data.append({
                'Professor': professor,
                'Total Courses': f"{total_courses}/{max_total}",
                'Total Utilization %': f"{(total_courses/max_total)*100:.1f}%" if max_total > 0 else "N/A"
            })
            
            for term in terms:
                streams_in_term = prof_load['streams_per_term'][term]
                max_streams = professor_term_limits.get((professor, term), 0)
                utilization = (streams_in_term/max_streams)*100 if max_streams > 0 else 0
                
                workload_data.append({
                    'Professor': f"  {term}",
                    'Total Courses': f"{streams_in_term}/{max_streams} streams",
                    'Total Utilization %': f"{utilization:.1f}%"
                })
        
        workload_df = pd.DataFrame(workload_data)
        st.dataframe(workload_df, hide_index=True)
        
        if solution.get('unassigned_offerings'):
            st.subheader("⚠️ Unassigned Course Offerings")
            for course, term in solution['unassigned_offerings']:
                streams = course_streams.get((course, term), 1)
                st.error(f"**{course}** in **{term}** ({streams} streams) - Could not assign")
        else:
            st.success("All course offerings successfully assigned!")
            
    elif solution['status'] == 'Infeasible':
        st.error("❌ Problem is infeasible - no solution exists")
        
        if solution.get('constraint_violations'):
            st.subheader("Possible Issues:")
            for violation in solution['constraint_violations']:
                st.write(f"• {violation}")
        
        st.subheader("Suggestions:")
        st.write("• Increase professor term limits (L_jk)")
        st.write("• Increase total course loads (b_j)")
        st.write("• Reduce number of course offerings")
        st.write("• Increase preference scores (avoid too many 0s)")
        
    else:
        st.error(f"Optimization failed: {solution['status']}")
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("← Back"):
            st.session_state.step = st.session_state.step - 1
            st.rerun()
    
    with col2:
        if st.button("Start Over"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()


def show_excel_upload_step():
    """Show Excel upload interface with specific sheet structure."""
    st.header("Excel File Upload")
    
    st.subheader("Required Excel File Format")
    st.markdown("""
    Your Excel file must contain exactly 4 sheets with this structure:
    
    **Sheet 1: c_ij (Course Preferences)**
    - Staff names in column A starting from A3
    - Course names in row 1 starting from B1
    - Course codes in row 2 starting from B2
    - Preference scores (0-10) in the data area
    
    **Sheet 2: t_jk, L_jk, b_j (Professor Constraints)**
    - Staff names in column A starting from A3
    - t_jk (Term Preferences): 3 columns for T1, T2, T3 (0-10 scale)
    - L_jk (Term Max Load): 3 columns for T1, T2, T3 (max streams per term)
    - b_j (Total Teaching Load): 1 column for total courses
    
    **Sheet 3: O_jk (Course Offerings)**
    - Course codes in column A
    - T1, T2, T3 columns with 1/0 values (1 = offered, 0 = not offered)
    
    **Sheet 4: n_jk (Course Streams)**
    - Course codes in column A  
    - T1, T2, T3 columns with stream counts (number of streams per course per term)
    """)
    
    with st.expander("Click to see sample Excel format"):
        col1, col2 = st.columns(2)
        
        with col1:
            st.write("**Sheet 1 - c_ij (Course Preferences):**")
            st.code("""
    A       B       C       D
1           ACTL1   ACTL2   ACTL3
2           Act1    Act2    Act3  
3   Jonathan  8       5       3
4   JK        6       9       4
5   Patrick   4       7       8
            """)
    
    if st.button("Download Excel Template"):
        template_data = create_excel_template_structured()
        if template_data:
            st.download_button(
                label="Download Template File",
                data=template_data,
                file_name="course_optimizer_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("Excel template generation requires additional libraries. Please create manually using the format above.")
    
    uploaded_file = st.file_uploader(
        "Upload your Excel file:",
        type=['xlsx', 'xls'],
        help="Upload an Excel file with the 4 required sheets"
    )
    
    if uploaded_file is not None:
        try:
            with st.spinner("Reading Excel file..."):
                excel_data = pd.ExcelFile(uploaded_file)
                
                required_sheets = ['c_ij', 't_jk_L_jk_b_j', 'O_jk', 'n_jk']
                available_sheets = excel_data.sheet_names
                
                st.write("Available sheets:", available_sheets)
                
                sheet_mapping = {}
                for req_sheet in required_sheets:
                    matched = False
                    for avail_sheet in available_sheets:
                        if req_sheet.lower() in avail_sheet.lower() or avail_sheet.lower() in req_sheet.lower():
                            sheet_mapping[req_sheet] = avail_sheet
                            matched = True
                            break
                    if not matched:
                        if req_sheet == 'c_ij' and len(available_sheets) > 0:
                            sheet_mapping[req_sheet] = available_sheets[0]
                        elif req_sheet == 't_jk_L_jk_b_j' and len(available_sheets) > 1:
                            sheet_mapping[req_sheet] = available_sheets[1]
                        elif req_sheet == 'O_jk' and len(available_sheets) > 2:
                            sheet_mapping[req_sheet] = available_sheets[2]
                        elif req_sheet == 'n_jk' and len(available_sheets) > 3:
                            sheet_mapping[req_sheet] = available_sheets[3]
                
                st.write("Sheet mapping:", sheet_mapping)
                
                if len(sheet_mapping) != 4:
                    st.error(f"Could not find all required sheets. Found: {list(sheet_mapping.values())}")
                    return
                
                sheet1 = pd.read_excel(uploaded_file, sheet_name=sheet_mapping['c_ij'], header=None)
                sheet2 = pd.read_excel(uploaded_file, sheet_name=sheet_mapping['t_jk_L_jk_b_j'])
                sheet3 = pd.read_excel(uploaded_file, sheet_name=sheet_mapping['O_jk'])
                sheet4 = pd.read_excel(uploaded_file, sheet_name=sheet_mapping['n_jk'])
                
                st.subheader("Processing Course Preferences (c_ij)")
                
                course_names = []
                course_codes = []
                for col_idx in range(1, len(sheet1.columns)):
                    if pd.notna(sheet1.iloc[0, col_idx]):
                        course_names.append(str(sheet1.iloc[0, col_idx]).strip())
                    if pd.notna(sheet1.iloc[1, col_idx]):
                        course_codes.append(str(sheet1.iloc[1, col_idx]).strip())
                
                staff_names = []
                for row_idx in range(2, len(sheet1)):
                    if pd.notna(sheet1.iloc[row_idx, 0]):
                        staff_names.append(str(sheet1.iloc[row_idx, 0]).strip())
                
                course_preferences = {}
                for staff_idx, staff in enumerate(staff_names):
                    for course_idx, course in enumerate(course_codes):
                        if course_idx < len(course_codes):
                            row_idx = staff_idx + 2
                            col_idx = course_idx + 1
                            if row_idx < len(sheet1) and col_idx < len(sheet1.columns):
                                pref_val = sheet1.iloc[row_idx, col_idx]
                                if pd.notna(pref_val):
                                    course_preferences[(course, staff)] = float(pref_val)
                                else:
                                    course_preferences[(course, staff)] = 0.0
                
                st.success(f"Found {len(staff_names)} staff and {len(course_codes)} courses")
                st.write("Staff:", staff_names)
                st.write("Courses:", course_codes)
                
                st.subheader("Processing Professor Constraints")
                
                terms = ['T1', 'T2', 'T3']
                term_preferences = {}
                professor_term_limits = {}
                professor_total_load = {}
                
                for idx, row in sheet2.iterrows():
                    if pd.notna(row.iloc[0]):
                        staff = str(row.iloc[0]).strip()
                        
                        for i, term in enumerate(terms):
                            if len(row) > i + 1 and pd.notna(row.iloc[i + 1]):
                                term_preferences[(staff, term)] = float(row.iloc[i + 1])
                            else:
                                term_preferences[(staff, term)] = 5.0
                        
                        for i, term in enumerate(terms):
                            if len(row) > i + 4 and pd.notna(row.iloc[i + 4]):
                                professor_term_limits[(staff, term)] = int(row.iloc[i + 4])
                            else:
                                professor_term_limits[(staff, term)] = 2
                        
                        if len(row) > 7 and pd.notna(row.iloc[7]):
                            professor_total_load[staff] = int(row.iloc[7])
                        else:
                            professor_total_load[staff] = 4
                
                st.subheader("Processing Course Offerings")
                
                course_offerings = {}
                for idx, row in sheet3.iterrows():
                    if pd.notna(row.iloc[0]):
                        course = str(row.iloc[0]).strip()
                        for i, term in enumerate(terms):
                            if len(row) > i + 1 and pd.notna(row.iloc[i + 1]):
                                course_offerings[(course, term)] = int(row.iloc[i + 1])
                            else:
                                course_offerings[(course, term)] = 0
                
                st.subheader("Processing Course Streams")
                
                course_streams = {}
                for idx, row in sheet4.iterrows():
                    if pd.notna(row.iloc[0]):
                        course = str(row.iloc[0]).strip()
                        for i, term in enumerate(terms):
                            if len(row) > i + 1 and pd.notna(row.iloc[i + 1]):
                                course_streams[(course, term)] = int(row.iloc[i + 1])
                            else:
                                course_streams[(course, term)] = 0
                
                st.session_state.courses = course_codes
                st.session_state.professors = staff_names
                st.session_state.terms = terms
                st.session_state.course_preferences = course_preferences
                st.session_state.term_preferences = term_preferences
                st.session_state.professor_term_limits = professor_term_limits
                st.session_state.professor_total_load = professor_total_load
                st.session_state.course_offerings = course_offerings
                st.session_state.course_streams = course_streams
                
                st.success("Excel file loaded successfully!")
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Courses", len(course_codes))
                with col2:
                    st.metric("Professors", len(staff_names))
                with col3:
                    offerings_count = sum(course_offerings.values())
                    st.metric("Course Offerings", offerings_count)
                
                st.subheader("Data Preview")
                
                with st.expander("Course Preferences (c_ij)"):
                    if course_preferences:
                        pref_matrix = pd.DataFrame(index=course_codes, columns=staff_names)
                        for course in course_codes:
                            for staff in staff_names:
                                pref_matrix.loc[course, staff] = course_preferences.get((course, staff), 0)
                        st.dataframe(pref_matrix)
                
                with st.expander("Professor Constraints"):
                    constraints_data = []
                    for staff in staff_names:
                        row_data = {'Professor': staff, 'Total Load (b_j)': professor_total_load.get(staff, 0)}
                        for term in terms:
                            row_data[f'{term} Pref (t_jk)'] = term_preferences.get((staff, term), 0)
                            row_data[f'{term} Limit (L_jk)'] = professor_term_limits.get((staff, term), 0)
                        constraints_data.append(row_data)
                    
                    constraints_df = pd.DataFrame(constraints_data)
                    st.dataframe(constraints_df)
                
                with st.expander("Course Offerings & Streams"):
                    offerings_data = []
                    for course in course_codes:
                        row_data = {'Course': course}
                        for term in terms:
                            offering = course_offerings.get((course, term), 0)
                            streams = course_streams.get((course, term), 0)
                            if offering == 1:
                                row_data[term] = f"✓ ({streams} streams)"
                            else:
                                row_data[term] = "✗"
                        offerings_data.append(row_data)
                    
                    offerings_df = pd.DataFrame(offerings_data)
                    st.dataframe(offerings_df)
                
                if st.button("Analyze Data First", type="primary"):
                    st.session_state.step = 2
                    st.rerun()
                    
        except Exception as e:
            st.error(f"Error reading Excel file: {str(e)}")
            st.write("Please check that your file format matches the required template.")
            st.code(str(e))


def show_data_analysis_step():
    """Show data analysis and statistics before optimization."""
    st.header("Step 2: Data Analysis & Statistics")
    
    courses = st.session_state.courses
    professors = st.session_state.professors
    terms = st.session_state.terms
    course_preferences = st.session_state.course_preferences
    term_preferences = st.session_state.term_preferences
    professor_term_limits = st.session_state.professor_term_limits
    professor_total_load = st.session_state.professor_total_load
    course_offerings = st.session_state.course_offerings
    course_streams = st.session_state.course_streams
    
    st.subheader("Data Overview")
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Courses", len(courses))
    with col2:
        st.metric("Professors", len(professors))
    with col3:
        offerings_count = sum(course_offerings.values())
        st.metric("Total Offerings", offerings_count)
    with col4:
        total_streams = sum(course_streams.values())
        st.metric("Total Streams", total_streams)
    
    st.subheader("Course Preferences Heatmap (c_ij)")
    st.markdown("**Scale: 0 = Cannot teach, 10 = Strongly prefer**")
    
    course_pref_matrix = pd.DataFrame(index=courses, columns=professors, dtype=float)
    
    for course in courses:
        for prof in professors:
            pref_value = course_preferences.get((course, prof), 0)
            course_pref_matrix.loc[course, prof] = pref_value
    
    fig1 = px.imshow(
        course_pref_matrix.values,
        labels=dict(x="Professor", y="Course", color="Preference Score"),
        x=course_pref_matrix.columns.tolist(),
        y=course_pref_matrix.index.tolist(),
        color_continuous_scale="RdYlGn",
        range_color=[0, 10],
        title="Course Preferences Matrix - All Courses vs All Staff",
        aspect="auto",
        text_auto=True
    )
    
    fig1.update_layout(
        height=max(800, len(courses) * 20),
        width=max(1200, len(professors) * 40),
        font=dict(size=10),
        xaxis=dict(tickangle=90, title="Staff Members"),
        yaxis=dict(
            title="Courses",
            tickmode='array',
            tickvals=list(range(len(courses))),
            ticktext=courses,
            tickfont=dict(size=9)
        )
    )
    
    st.plotly_chart(fig1, use_container_width=True)
    
    col1, col2, col3 = st.columns(3)
    with col1:
        avg_pref = course_pref_matrix.mean().mean()
        st.metric("Average Preference", f"{avg_pref:.2f}")
    with col2:
        high_prefs = (course_pref_matrix >= 8).sum().sum()
        st.metric("High Preferences (≥8)", high_prefs)
    with col3:
        pref_range = f"{course_pref_matrix.min().min():.0f} - {course_pref_matrix.max().max():.0f}"
        st.metric("Preference Range", pref_range)
    
    st.subheader("Hierarchical Clustering Analysis")
    
    clustering_type = st.radio(
        "Select clustering view:",
        ["Course Clustering (Which courses are similar?)", 
         "Professor Clustering (Which professors have similar preferences?)"],
        horizontal=True
    )
    
    # Add PCA section - with optional sklearn
    try:
        from sklearn.decomposition import PCA
        from sklearn.cluster import KMeans
        from sklearn.preprocessing import StandardScaler
        
        st.subheader("PCA & K-Means Clustering Visualization")
        st.markdown("**Visualize data in 2D using Principal Component Analysis with K-Means clustering**")
        
        # Number of clusters selector
        n_clusters = st.slider(
            "Select number of clusters (K):",
            min_value=2,
            max_value=5,
            value=3,
            help="Choose how many clusters to identify in the data"
        )
        
        if clustering_type == "Course Clustering (Which courses are similar?)":
            st.markdown(f"**PCA plot of courses with {n_clusters} K-Means clusters**")
            
            # Prepare data: courses as observations (rows), professors as features (columns)
            data_for_pca = course_pref_matrix.values
            
            if len(courses) >= 2:
                # Standardize the data
                scaler = StandardScaler()
                data_scaled = scaler.fit_transform(data_for_pca)
                
                # Apply PCA
                pca = PCA(n_components=2)
                pca_result = pca.fit_transform(data_scaled)
                
                # Apply K-Means clustering
                kmeans = KMeans(n_clusters=n_clusters, random_state=42, n_init=10)
                cluster_labels = kmeans.fit_predict(data_scaled)
                
                # Create DataFrame for plotting
                pca_df = pd.DataFrame({
                    'PC1': pca_result[:, 0],
                    'PC2': pca_result[:, 1],
                    'Course': courses,
                    'Cluster': [f'Cluster {i+1}' for i in cluster_labels]
                })
                
                # Create PCA scatter plot
                fig_pca = px.scatter(
                    pca_df,
                    x='PC1',
                    y='PC2',
                    color='Cluster',
                    text='Course',
                    title=f'PCA of Courses with {n_clusters} K-Means Clusters',
                    labels={'PC1': f'PC1 ({pca.explained_variance_ratio_[0]:.1%} variance)',
                            'PC2': f'PC2 ({pca.explained_variance_ratio_[1]:.1%} variance)'},
                    color_discrete_sequence=px.colors.qualitative.Set2
                )
                
                fig_pca.update_traces(
                    textposition='top center',
                    marker=dict(size=12, line=dict(width=2, color='white'))
                )
                
                fig_pca.update_layout(
                    height=700,
                    width=1000,
                    font=dict(size=12),
                    showlegend=True
                )
                
                st.plotly_chart(fig_pca, use_container_width=True)
                
                # Show cluster membership
                st.markdown("**Cluster Membership:**")
                cluster_summary = []
                for i in range(n_clusters):
                    cluster_courses = pca_df[pca_df['Cluster'] == f'Cluster {i+1}']['Course'].tolist()
                    cluster_summary.append({
                        'Cluster': f'Cluster {i+1}',
                        'Size': len(cluster_courses),
                        'Courses': ', '.join(cluster_courses)
                    })
                
                cluster_df = pd.DataFrame(cluster_summary)
                st.dataframe(cluster_df, hide_index=True, use_container_width=True)
                
                # Variance explained
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("PC1 Variance Explained", f"{pca.explained_variance_ratio_[0]:.1%}")
                with col2:
                    st.metric("PC2 Variance Explained", f"{pca.explained_variance_ratio_[1]:.1%}")
                with col3:
                    st.metric("Total Variance Explained", f"{sum(pca.explained_variance_ratio_):.1%}")
            else:
                st.warning("Need at least 2 courses for PCA analysis")
        
        else:  # Professor clustering
            st.markdown(f"**PCA plot of professors with {n_clusters} K-Means clusters**")
            
            # Prepare data: professors as observations (rows), courses as features (columns)
            data_for_pca = course_pref_matrix.T.values
            
            if len(professors) >= 2:
                # Standardize the data
                scaler = StandardScaler()
                data_scaled = scaler.fit_transform(data_for_pca)
                
                # Apply PCA
                pca = PCA(n_components=2)
                pca_result = pca.fit_transform(data_scaled)
                
                # Apply K-Means clustering
                kmeans = KMeans(n_clusters=n_clusters, random_state=42, n_init=10)
                cluster_labels = kmeans.fit_predict(data_scaled)
                
                # Create DataFrame for plotting
                pca_df = pd.DataFrame({
                    'PC1': pca_result[:, 0],
                    'PC2': pca_result[:, 1],
                    'Professor': professors,
                    'Cluster': [f'Cluster {i+1}' for i in cluster_labels]
                })
                
                # Create PCA scatter plot
                fig_pca = px.scatter(
                    pca_df,
                    x='PC1',
                    y='PC2',
                    color='Cluster',
                    text='Professor',
                    title=f'PCA of Professors with {n_clusters} K-Means Clusters',
                    labels={'PC1': f'PC1 ({pca.explained_variance_ratio_[0]:.1%} variance)',
                            'PC2': f'PC2 ({pca.explained_variance_ratio_[1]:.1%} variance)'},
                    color_discrete_sequence=px.colors.qualitative.Set2
                )
                
                fig_pca.update_traces(
                    textposition='top center',
                    marker=dict(size=12, line=dict(width=2, color='white'))
                )
                
                fig_pca.update_layout(
                    height=700,
                    width=1000,
                    font=dict(size=12),
                    showlegend=True
                )
                
                st.plotly_chart(fig_pca, use_container_width=True)
                
                # Show cluster membership
                st.markdown("**Cluster Membership:**")
                cluster_summary = []
                for i in range(n_clusters):
                    cluster_profs = pca_df[pca_df['Cluster'] == f'Cluster {i+1}']['Professor'].tolist()
                    cluster_summary.append({
                        'Cluster': f'Cluster {i+1}',
                        'Size': len(cluster_profs),
                        'Professors': ', '.join(cluster_profs)
                    })
                
                cluster_df = pd.DataFrame(cluster_summary)
                st.dataframe(cluster_df, hide_index=True, use_container_width=True)
                
                # Variance explained
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("PC1 Variance Explained", f"{pca.explained_variance_ratio_[0]:.1%}")
                with col2:
                    st.metric("PC2 Variance Explained", f"{pca.explained_variance_ratio_[1]:.1%}")
                with col3:
                    st.metric("Total Variance Explained", f"{sum(pca.explained_variance_ratio_):.1%}")
            else:
                st.warning("Need at least 2 professors for PCA analysis")
        
        st.markdown("---")
        
    except ImportError:
        st.info("💡 **PCA & K-Means Clustering not available**. Install scikit-learn to enable this feature: `pip install scikit-learn`")
    
    st.subheader("Hierarchical Clustering Dendrograms")
    st.markdown("**View hierarchical relationships in the data**")
    
    if clustering_type == "Course Clustering (Which courses are similar?)":
        st.markdown("**Clusters courses based on which professors prefer them similarly**")
        
        # Course preferences: courses as rows, professors as columns
        # For clustering courses, we need courses as observations (rows)
        course_matrix_for_clustering = course_pref_matrix  # Already has courses as rows
        
        # Check if we have enough courses for clustering
        if len(courses) < 2:
            st.warning("Need at least 2 courses for clustering analysis")
        else:
            # Calculate linkage using the course matrix directly
            linkage_matrix = linkage(course_matrix_for_clustering.values, method='ward')
            
            fig_dendro = go.Figure()
            
            dendro_data = dendrogram(linkage_matrix, labels=courses, no_plot=True)
            
            icoord = np.array(dendro_data['icoord'])
            dcoord = np.array(dendro_data['dcoord'])
            
            for i in range(len(icoord)):
                fig_dendro.add_trace(go.Scatter(
                    x=icoord[i],
                    y=dcoord[i],
                    mode='lines',
                    line=dict(color='rgb(100,100,100)', width=1),
                    showlegend=False,
                    hoverinfo='skip'
                ))
            
            fig_dendro.update_layout(
                title="Course Dendrogram - Hierarchical Clustering",
                xaxis=dict(
                    title="Courses",
                    tickvals=list(range(5, len(courses)*10+5, 10)),
                    ticktext=dendro_data['ivl'],
                    tickangle=90
                ),
                yaxis=dict(title="Distance (Ward Linkage)"),
                height=600,
                showlegend=False
            )
            
            st.plotly_chart(fig_dendro, use_container_width=True)
            
            st.markdown("**Clustered Heatmap (Courses reordered by similarity)**")
            
            course_order = [courses[i] for i in dendro_data['leaves']]
            clustered_matrix = course_pref_matrix.loc[course_order, :]
            
            fig_clustered = px.imshow(
                clustered_matrix.values,
                labels=dict(x="Professor", y="Course", color="Preference Score"),
                x=clustered_matrix.columns.tolist(),
                y=clustered_matrix.index.tolist(),
                color_continuous_scale="RdYlGn",
                range_color=[0, 10],
                title="Clustered Course Preferences (Courses Reordered)",
                aspect="auto",
                text_auto=True
            )
            
            fig_clustered.update_layout(
                height=max(800, len(courses) * 20),
                width=max(1200, len(professors) * 40),
                font=dict(size=10),
                xaxis=dict(tickangle=90, title="Staff Members"),
                yaxis=dict(
                    title="Courses (Clustered Order)",
                    tickmode='array',
                    tickvals=list(range(len(course_order))),
                    ticktext=course_order,
                    tickfont=dict(size=9)
                )
            )
            
            st.plotly_chart(fig_clustered, use_container_width=True)
        
    else:
        st.markdown("**Clusters professors based on their course preference patterns**")
        
        # For clustering professors, we need professors as observations (rows)
        # Transpose so professors are rows
        prof_matrix_for_clustering = course_pref_matrix.T  # Now professors are rows
        
        # Check if we have enough professors for clustering
        if len(professors) < 2:
            st.warning("Need at least 2 professors for clustering analysis")
        else:
            # Calculate linkage using professor matrix
            linkage_matrix = linkage(prof_matrix_for_clustering.values, method='ward')
            
            fig_dendro_prof = go.Figure()
            
            dendro_data_prof = dendrogram(linkage_matrix, labels=professors, no_plot=True)
            
            icoord = np.array(dendro_data_prof['icoord'])
            dcoord = np.array(dendro_data_prof['dcoord'])
            
            for i in range(len(icoord)):
                fig_dendro_prof.add_trace(go.Scatter(
                    x=icoord[i],
                    y=dcoord[i],
                    mode='lines',
                    line=dict(color='rgb(100,100,100)', width=1),
                    showlegend=False,
                    hoverinfo='skip'
                ))
            
            fig_dendro_prof.update_layout(
                title="Professor Dendrogram - Hierarchical Clustering",
                xaxis=dict(
                    title="Professors",
                    tickvals=list(range(5, len(professors)*10+5, 10)),
                    ticktext=dendro_data_prof['ivl'],
                    tickangle=90
                ),
                yaxis=dict(title="Distance (Ward Linkage)"),
                height=600,
                showlegend=False
            )
            
            st.plotly_chart(fig_dendro_prof, use_container_width=True)
            
            st.markdown("**Clustered Heatmap (Professors reordered by similarity)**")
            
            prof_order = [professors[i] for i in dendro_data_prof['leaves']]
            clustered_matrix_prof = course_pref_matrix.loc[:, prof_order]
            
            fig_clustered_prof = px.imshow(
                clustered_matrix_prof.values,
                labels=dict(x="Professor", y="Course", color="Preference Score"),
                x=clustered_matrix_prof.columns.tolist(),
                y=clustered_matrix_prof.index.tolist(),
                color_continuous_scale="RdYlGn",
                range_color=[0, 10],
                title="Clustered Course Preferences (Professors Reordered)",
                aspect="auto",
                text_auto=True
            )
            
            fig_clustered_prof.update_layout(
                height=max(800, len(courses) * 20),
                width=max(1200, len(prof_order) * 40),
                font=dict(size=10),
                xaxis=dict(
                    tickangle=90, 
                    title="Staff Members (Clustered Order)",
                    tickmode='array',
                    tickvals=list(range(len(prof_order))),
                    ticktext=prof_order
                ),
                yaxis=dict(
                    title="Courses",
                    tickmode='array',
                    tickvals=list(range(len(courses))),
                    ticktext=courses,
                    tickfont=dict(size=9)
                )
            )
            
            st.plotly_chart(fig_clustered_prof, use_container_width=True)
    
    st.subheader("Course Preference Analysis")
    
    all_scores = set()
    for course in courses:
        for prof in professors:
            score = course_pref_matrix.loc[course, prof]
            if score > 0:
                all_scores.add(int(score))
    
    sorted_scores = sorted(all_scores, reverse=True)
    
    scores_with_conflicts = []
    all_score_data = {}
    
    for score in sorted_scores:
        courses_with_multiple_score = []
        for course in courses:
            profs_with_score = []
            for prof in professors:
                prof_score = int(course_pref_matrix.loc[course, prof])
                if prof_score == score:
                    profs_with_score.append(prof)
            
            if len(profs_with_score) > 1:
                course_terms = []
                for term in terms:
                    if course_offerings.get((course, term), 0) == 1:
                        course_terms.append(term)
                
                terms_offered = ', '.join(course_terms) if course_terms else 'Not offered'
                
                courses_with_multiple_score.append({
                    'Course': course,
                    'Terms Offered': terms_offered,
                    'Professors': ', '.join(profs_with_score),
                    'Count': len(profs_with_score)
                })
        
        if courses_with_multiple_score:
            scores_with_conflicts.append(score)
            all_score_data[score] = courses_with_multiple_score
    
    if scores_with_conflicts:
        tab_names = [f"Score {score}" for score in scores_with_conflicts]
        
        score_tabs = st.tabs(tab_names)
        
        for i, score in enumerate(scores_with_conflicts):
            with score_tabs[i]:
                st.markdown(f"**Courses where multiple professors gave score {score}:**")
                
                score_df = pd.DataFrame(all_score_data[score])
                st.dataframe(score_df, hide_index=True, use_container_width=True)
                
                total_conflicts = len(all_score_data[score])
                total_professors = sum([row['Count'] for row in all_score_data[score]])
                
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Courses with Conflicts", total_conflicts)
                with col2:
                    st.metric("Total Professor Instances", total_professors)
    else:
        st.info("No courses have multiple professors with the same score")
    
    st.subheader("Term Preferences Heatmap (t_jk)")
    st.markdown("**Scale: 0 = Cannot teach, 10 = Strongly prefer**")
    
    term_pref_matrix = pd.DataFrame(index=professors, columns=terms, dtype=float)
    
    for prof in professors:
        for term in terms:
            term_value = term_preferences.get((prof, term), 0)
            term_pref_matrix.loc[prof, term] = term_value
    
    fig2 = px.imshow(
        term_pref_matrix.values,
        labels=dict(x="Term", y="Professor", color="Preference Score"),
        x=term_pref_matrix.columns.tolist(),
        y=term_pref_matrix.index.tolist(),
        color_continuous_scale="RdYlGn",
        range_color=[0, 10],
        title="Term Preferences Matrix - All Staff vs Terms",
        aspect="auto",
        text_auto=True
    )
    
    fig2.update_layout(
        height=max(600, len(professors) * 25),
        width=800,
        font=dict(size=12),
        xaxis=dict(title="Terms"),
        yaxis=dict(title="Staff Members")
    )
    
    st.plotly_chart(fig2, use_container_width=True)
    
    col1, col2, col3 = st.columns(3)
    with col1:
        avg_term_pref = term_pref_matrix.mean().mean()
        st.metric("Average Term Preference", f"{avg_term_pref:.2f}")
    with col2:
        term_avg = term_pref_matrix.mean()
        most_popular_term = term_avg.idxmax()
        st.metric("Most Popular Term", f"{most_popular_term} ({term_avg[most_popular_term]:.1f})")
    with col3:
        term_range = f"{term_pref_matrix.min().min():.0f} - {term_pref_matrix.max().max():.0f}"
        st.metric("Term Preference Range", term_range)
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("← Back to Upload"):
            st.session_state.step = 1
            st.rerun()
    
    with col2:
        if st.button("Run Optimization →", type="primary"):
            st.session_state.step = 3
            st.rerun()


def create_excel_template_structured():
    """Create an Excel template file with the specified structure."""
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        import io
        
        wb = Workbook()
        
        ws1 = wb.active
        ws1.title = "c_ij"
        
        course_names = ["ACTL1", "ACTL2", "ACTL3", "ACTL4", "ACTL5"]
        course_codes = ["Act1", "Act2", "Act3", "Act4", "Act5"]
        staff_names = ["Jonathan", "JK", "Patrick", "Andres"]
        
        for i, name in enumerate(course_names, start=2):
            ws1.cell(row=1, column=i, value=name)
        
        for i, code in enumerate(course_codes, start=2):
            ws1.cell(row=2, column=i, value=code)
        
        for i, staff in enumerate(staff_names, start=3):
            ws1.cell(row=i, column=1, value=staff)
        
        for staff_idx in range(len(staff_names)):
            for course_idx in range(len(course_codes)):
                ws1.cell(row=staff_idx+3, column=course_idx+2, value=5)
        
        ws2 = wb.create_sheet("t_jk_L_jk_b_j")
        
        headers = ["Staff", "T1", "T2", "T3", "T1", "T2", "T3", "Total"]
        subheaders = ["", "Term Prefs", "", "", "Max Streams", "", "", "Load"]
        
        for i, header in enumerate(headers, start=1):
            ws2.cell(row=1, column=i, value=header)
        for i, subheader in enumerate(subheaders, start=1):
            ws2.cell(row=2, column=i, value=subheader)
        
        for i, staff in enumerate(staff_names, start=3):
            ws2.cell(row=i, column=1, value=staff)
            for j in range(2, 5):
                ws2.cell(row=i, column=j, value=5)
            for j in range(5, 8):
                ws2.cell(row=i, column=j, value=2)
            ws2.cell(row=i, column=8, value=4)
        
        ws3 = wb.create_sheet("O_jk")
        
        ws3.cell(row=1, column=1, value="Course")
        terms = ["T1", "T2", "T3"]
        for i, term in enumerate(terms, start=2):
            ws3.cell(row=1, column=i, value=term)
        
        for i, course in enumerate(course_codes, start=2):
            ws3.cell(row=i, column=1, value=course)
            ws3.cell(row=i, column=2, value=1 if i % 2 == 0 else 0)
            ws3.cell(row=i, column=3, value=1 if i % 3 == 0 else 0)
            ws3.cell(row=i, column=4, value=1)
        
        ws4 = wb.create_sheet("n_jk")
        
        ws4.cell(row=1, column=1, value="Course")
        for i, term in enumerate(terms, start=2):
            ws4.cell(row=1, column=i, value=term)
        
        for i, course in enumerate(course_codes, start=2):
            ws4.cell(row=i, column=1, value=course)
            for j in range(2, 5):
                offering = ws3.cell(row=i, column=j).value
                if offering == 1:
                    ws4.cell(row=i, column=j, value=1)
                else:
                    ws4.cell(row=i, column=j, value=0)
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
        
    except ImportError:
        return None
    except Exception as e:
        st.error(f"Error creating template: {str(e)}")
        return None


if __name__ == "__main__":
    main()

            
